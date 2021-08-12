try:
    from BeautifulSoup import BeautifulSoup
except ImportError:
    from bs4 import BeautifulSoup
import requests
import urllib.request
import urllib
import json
import re
import openpyxl
import datetime
import pymysql


headers = {
    'accept-encoding': 'deflate',
    'accept-language': 'zh-CN,zh;q=0.9',
    'pragma': 'no-cache',
    'cache-control': 'no-cache',
    'upgrade-insecure-requests': '1',
    'user-agent': "Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1",
}

HEADERS = {'user-agent': "Mozilla/5.0 (iPhone; CPU iPhone OS 11_0 like Mac OS X) AppleWebKit/604.1.38 (KHTML, like Gecko) Version/11.0 Mobile/15A372 Safari/604.1"}

mapCode2Name = {"0xe602":"num_","0xe605":"num_3","0xe606":"num_4","0xe603":"num_1","0xe604":"num_2","0xe618":"num_","0xe619":"num_4","0xe60a":"num_8","0xe60b":"num_9","0xe60e":"num_","0xe60f":"num_5","0xe60c":"num_4", \
                "0xe60d":"num_1","0xe612":"num_6","0xe613":"num_8","0xe610":"num_3","0xe611":"num_2","0xe616":"num_1","0xe617":"num_3","0xe614":"num_9","0xe615":"num_7","0xe609":"num_7","0xe607":"num_5","0xe608":"num_6","0xe61b":"num_5", \
                "0xe61c":"num_8","0xe61a":"num_2","0xe61f":"num_6","0xe61d":"num_9","0xe61e":"num_7"}
mapCode2Font = {"num_9":8,"num_5":5,"num_6":6,"num_":1,"num_7":9,"num_8":7,"num_1":0,"num_2":3,"num_3":2,"num_4":4}

def getUserInfo(shared_url, **headers):
    real_url = getRealAddress(shared_url)
    parsed = urllib.parse.urlparse(real_url)
    hostname = parsed.hostname
    sec_uid = urllib.parse.parse_qs(parsed.query)['sec_uid']
    user_info_url = "https://%s/web/api/v2/user/info/" % hostname
    user_info_params = { 'sec_uid': sec_uid }
    res = requests.get(user_info_url, headers=headers,
                       params=user_info_params).json()
    user_info = res['user_info']
    user_avatar = user_info['avatar_larger']['url_list'][2]
    user_nickname = user_info['nickname']
    user_sign = user_info['signature']
    user_id = user_info['unique_id']
    count_of_videos = user_info['aweme_count']
    follower_count = user_info['follower_count']
    following_count = user_info['following_count']
    zan_count = user_info['total_favorited']
    like_count = user_info['favoriting_count']
    return {'user_avatar':user_avatar, 'user_nickname':user_nickname, 'user_sign':user_sign, 'user_id':user_id,     #'user_avatar':user_avatar,
            'count_of_videos':count_of_videos, 'follower_count':follower_count, 'following_count':following_count,
            'zan_count':zan_count, 'like_count':like_count, 'like_count':like_count}

def getUserVideos(url):
    number = re.findall(r'share/user/(\d+)', url)
    if not len(number):
        return
    dytk = get_dytk(url)
    # hostname = urllib.parse.urlparse(url).hostname
    # if hostname != 't.tiktok.com' and not dytk:
    #     return
    user_id = number[0]
    return getUserMedia(user_id, dytk, url)


def getRealAddress(url):
    if url.find('v.douyin.com') < 0:
        return url
    res = requests.get(url, headers=headers, allow_redirects=False)
    return res.headers['Location'] if res.status_code == 302 else None


def get_dytk(url):
    res = requests.get(url, headers=headers)
    if not res:
        return None
    dytk = re.findall("dytk: '(.*)'", res.content.decode('utf-8'))
    if len(dytk):
        return dytk[0]
    return None

def getUserMedia(user_id, dytk, url):
    videos = []
    parsed = urllib.parse.urlparse(url)
    hostname = parsed.hostname
    sec_uid = urllib.parse.parse_qs(parsed.query)['sec_uid']

    #signature = generateSignature(str(user_id))
    user_video_url = "https://%s/web/api/v2/aweme/post/" % hostname
    user_video_params = {
        'sec_uid': sec_uid,
        'count': '21',
        'max_cursor': '0',
        'aid': '1128',
        '_signature': '2Vx9mxAZh0o-K4Wdv7NFKNlcfY',
        'dytk': dytk
    }
    if hostname == 't.tiktok.com':
        user_video_params.pop('dytk')
        user_video_params['aid'] = '1180'

    max_cursor, video_count = None, 0
    while True:
        if max_cursor:
            user_video_params['max_cursor'] = str(max_cursor)
        res = requests.get(user_video_url, headers=headers,
                           params=user_video_params)
        contentJson = json.loads(res.content.decode('utf-8'))
        aweme_list = contentJson.get('aweme_list', [])
        for aweme in aweme_list:
            video_count += 1
            aweme['hostname'] = hostname
            video =  {
                'addr': aweme['video']['download_addr']['url_list'][0],      #  注意 play_addr /  download_addr
                'desc': aweme['desc'],
                'duration': aweme['video']['duration'],
                'cover': aweme['video']['cover']['url_list'][0],
                'statistics': aweme['statistics']
            }
            videos.append(video)
        if contentJson.get('has_more'):
            max_cursor = contentJson.get('max_cursor')
        else:
            break


    if video_count == 0:
        print("There's no video in number %s." % user_id)

    return videos



def getHtml(url,**headers):
    try:
        req = urllib.request.Request(url,headers=headers)
        resp = urllib.request.urlopen(req)
        return str(resp.read(), 'utf-8')
    except urllib.error.HTTPError as e:
        print(e.msg)
        return ''



def woff2tff(ls):
    res = ''
    for s in ls:
        res = res + formatNum(s)
    return res

def splitByChinese(s):
    p = re.compile("[\u4e00-\u9fa5]", re.U)
    return p.split(s)

def isChinese(s):
    p = re.compile("[\u4e00-\u9fa5]", re.U)
    result = p.match(s)
    if result :
        return True
    return False


def formatNum(s):
    if isChinese(s):
        return ''
    if len(s)<8 or s.find("hzsdxe6") < 0 :
        return s
    s1 = '0'+s[4:-1]
    res = mapCode2Font[mapCode2Name[s1]]
    return str(res)


def getUserAll(shared_url):
    profile = getUserInfo(shared_url, **HEADERS)
    if profile:
        videos = getUserVideos(getRealAddress(shared_url))
        profile['videos'] = videos
    return profile

def to_excel(userinfo):
    data = []
    sheet_columns = ["时间","用户名","用户id","粉丝","关注","点赞","作品","喜欢"]
    # json.loads(userinfo)
    today_time = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")

    follower_count = userinfo["follower_count"]
    following_count = userinfo["following_count"]
    zan_count = userinfo["zan_count"]
    user_id = userinfo["user_id"]
    user_nickname = userinfo["user_nickname"]
    count_of_videos = userinfo["count_of_videos"]
    like_count = userinfo["like_count"]
    data.append(today_time)
    data.append(user_nickname)
    data.append(user_id)
    data.append(follower_count)
    data.append(following_count)
    data.append(zan_count)
    data.append(count_of_videos)
    data.append(like_count)
    # print(data)
    wb = openpyxl.load_workbook(r"D:\Work\spider_data\douyin_spider_data" + "\\" +"douyin_spider" + ".xlsx")
    ws = wb.active
    ws.title = "user_info"
    ws.append(data)
    # ws.append(sheet_columns)
    # for i in range(len(data)):  #7
        # ws.cell(1, i+1).value = sheet_columns[i]

        # ws.cell(2, i+1).value = data[i]
    wb.save(r"D:\Work\spider_data\douyin_spider_data" + "\\" +"douyin_spider" + ".xlsx")
    print("done")


    # df = pd.DataFrame()
    # for line in range(sheet_columns):
    #     df1 = df.append(sheet_columns[line])
    #     print(df1)

    # workbook = xlwt.Workbook("D:\Work\spider_data\douyin_spider_data")
    # sheet1 = workbook.add_sheet('user_info')
    # # for i in range(len(userinfo) - 3):
    # #     sheet1.write(1, i+1, sheet_columns[i])
    # #     sheet1.write(2, data[i])
    # sheet1.write(1,1,"guo")


    # with open('D:\Work\spider_data\douyin_spider_data', 'r', encoding='UTF-8') as f:
    #     for line in f:

def data_to_db(userinfo):

    # data = []
    # sheet_columns = ["时间","用户名","用户id","粉丝","关注","点赞","作品","喜欢"]
    # json.loads(userinfo)
    # today_time = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")

    follower_count = userinfo["follower_count"]
    following_count = userinfo["following_count"]
    zan_count = userinfo["zan_count"]
    user_id = userinfo["user_id"]
    user_nickname = userinfo["user_nickname"]
    count_of_videos = userinfo["count_of_videos"]
    like_count = userinfo["like_count"]
    # data.append(today_time)
    # data.append(user_nickname)
    # data.append(user_id)
    # data.append(follower_count)
    # data.append(following_count)
    # data.append(zan_count)
    # data.append(count_of_videos)
    # data.append(like_count)

    # 建立数据库连接
    db = pymysql.connect(host="192.168.50.182",user="local_gsc",password="gsc123",database="jj_database" )
    cursor = db.cursor()
    sql = "INSERT INTO gj_mcn_spider_account_info(user_name, user_id, fans_count,follow_count, " \
          "zan_count,composition_count, like_count) values({},{},{},{},{},{}," \
          "{})".format('\'' + user_nickname + '\'','\'' + user_id + '\'',follower_count,following_count,zan_count,count_of_videos,like_count)
    cursor.execute(sql)
    db.commit()
    cursor.close()
    print("db done")

def videos_first_to_excel(userinfo):
    data = []
    user_data = []
    videos_data = []
    sheet_columns = ["时间","用户名","用户id","粉丝","关注","点赞","作品","喜欢","视频描述","视频时长","视频封面","视频转发","视频评论","视频点赞","aweme_id"]
    today_time = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    user_nickname = userinfo["user_nickname"]
    user_id = userinfo["user_id"]
    follower_count = userinfo["follower_count"]
    following_count = userinfo["following_count"]
    zan_count = userinfo["zan_count"]
    count_of_videos = userinfo["count_of_videos"]
    like_count = userinfo["like_count"]
    videos = userinfo["videos"]

    user_data.append(today_time)
    user_data.append(user_nickname)
    user_data.append(user_id)
    user_data.append(follower_count)
    user_data.append(following_count)
    user_data.append(zan_count)
    user_data.append(count_of_videos)
    user_data.append(like_count)


    wb = openpyxl.Workbook()    #(r"D:\Work\spider_data\douyin_spider_data" + "\\" +"douyin_spider" + ".xlsx")
    ws = wb.active
    ws.title = "videos_info"
    ws.append(sheet_columns)
    for i in range(len(userinfo["videos"])):

        # ws.append(user_data)
        data = []
        videos_data = []
        desc = userinfo["videos"][i]["desc"]
        duration = userinfo["videos"][i]["duration"]
        cover = userinfo["videos"][i]["cover"]
        share_count = userinfo["videos"][i]["statistics"]["share_count"]
        comment_count = userinfo["videos"][i]["statistics"]["comment_count"]
        digg_count = userinfo["videos"][i]["statistics"]["digg_count"]
        aweme_id = userinfo["videos"][i]["statistics"]["aweme_id"]
        videos_data.append(desc)
        videos_data.append(duration)
        videos_data.append(cover)
        videos_data.append(share_count)
        videos_data.append(comment_count)
        videos_data.append(digg_count)
        videos_data.append(aweme_id)
        data.extend(user_data)
        data.extend(videos_data)
        # print(data)
        # user_data.extend(videos_data)
    # print(user_data)
        # data.append(user_data)
        # data.append(videos_data)
        ws.append(data)
    # ws.append(sheet_columns)
    # for i in range(len(data)):  #7
    # ws.cell(1, i+1).value = sheet_columns[i]

    # ws.cell(2, i+1).value = data[i]
    # print(user_data)
    wb.save(r"D:\Work\spider_data\douyin_spider_data\douyin 视频信息spider.xlsx")
    print("first time write done to excel")
    # wb.save(r"D:\Work\spider_data\douyin_videos_spider_data" + "\\" +today_time + " douyin用户信息spider.xlsx")
    # r"D:\Work\spider_data\douyin_videos_spider_data" + "\\" +today_time + " douyin用户信息spider.xlsx"

def videos_to_excel(userinfo):
    data = []
    user_data = []
    videos_data = []
    sheet_columns = ["时间","用户名","用户id","粉丝","关注","点赞","作品","喜欢","视频描述","视频时长","视频封面","视频转发","视频评论","视频点赞","aweme_id"]
    today_time = datetime.datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    user_nickname = userinfo["user_nickname"]
    user_id = userinfo["user_id"]
    follower_count = userinfo["follower_count"]
    following_count = userinfo["following_count"]
    zan_count = userinfo["zan_count"]
    count_of_videos = userinfo["count_of_videos"]
    like_count = userinfo["like_count"]
    videos = userinfo["videos"]

    user_data.append(today_time)
    user_data.append(user_nickname)
    user_data.append(user_id)
    user_data.append(follower_count)
    user_data.append(following_count)
    user_data.append(zan_count)
    user_data.append(count_of_videos)
    user_data.append(like_count)


    wb = openpyxl.load_workbook(r"D:\Work\spider_data\douyin_spider_data\douyin 视频信息spider 3.xlsx")    #(r"D:\Work\spider_data\douyin_spider_data" + "\\" +"douyin_spider" + ".xlsx")
    ws = wb.active
    # ws.title = "videos_info"
    # ws.append(sheet_columns)
    for i in range(len(userinfo["videos"])):

        # ws.append(user_data)
        data = []
        videos_data = []
        desc = userinfo["videos"][i]["desc"]
        addr = userinfo["videos"][i]["addr"]
        duration = userinfo["videos"][i]["duration"]
        cover = userinfo["videos"][i]["cover"]
        share_count = userinfo["videos"][i]["statistics"]["share_count"]
        comment_count = userinfo["videos"][i]["statistics"]["comment_count"]
        digg_count = userinfo["videos"][i]["statistics"]["digg_count"]
        aweme_id = userinfo["videos"][i]["statistics"]["aweme_id"]
        videos_data.append(desc)
        videos_data.append(addr)
        videos_data.append(duration)
        videos_data.append(cover)
        videos_data.append(share_count)
        videos_data.append(comment_count)
        videos_data.append(digg_count)
        videos_data.append(aweme_id)
        data.extend(user_data)
        data.extend(videos_data)
        # print(data)
        # user_data.extend(videos_data)
        # print(user_data)
        # data.append(user_data)
        # data.append(videos_data)
        ws.append(data)
    # ws.append(sheet_columns)
    # for i in range(len(data)):  #7
    # ws.cell(1, i+1).value = sheet_columns[i]

    # ws.cell(2, i+1).value = data[i]
    # print(user_data)
    wb.save(r"D:\Work\spider_data\douyin_spider_data\douyin 视频信息spider 3.xlsx")

    print("After first time write done to excel")



def videos_to_db():
    pass



if __name__ == '__main__':
    # userInfo = getUserAll("https://v.douyin.com/eUf1dNh")   # 乔哥很耿直
    # print(userInfo)
    # userInfo1 = getUserAll("https://v.douyin.com/e5kkEqV")  # 百草讲坛
    # userInfo2 = getUserAll("https://v.douyin.com/eQj5LJX")  # 百草讲坛
    # videos_first_to_excel(userInfo)
    # videos_to_excel(userInfo2)
    #
    # userInfo = getUserAll("https://v.douyin.com/eUf1dNh")# 乔哥很耿直
    # userInfo1 = getUserAll("https://v.douyin.com/e5kkEqV")# 百草讲坛
    # userInfo2 = getUserAll("https://v.douyin.com/emHhqcB")  #蔷小草的田园生活
    # userInfo3 = getUserAll("https://v.douyin.com/emHLh2n")  #大贤和奶奶
    # userInfo4 = getUserAll("https://v.douyin.com/emHJkSo")  #本草中国
    # userInfo5 = getUserAll("https://v.douyin.com/emHMYHu")  #苏叶和京墨
    # userInfo6 = getUserAll("https://v.douyin.com/exKnyoG/")  #王碧瑶要做辟谣王

    #
    userInfo1 = getUserAll("https://v.douyin.com/eXYjYgS/")  #只穿高跟鞋的汪奶奶
    # print(userInfo1)
    userInfo2 = getUserAll("https://v.douyin.com/eXF79Lc/")  #末那大叔
    userInfo3 = getUserAll("https://v.douyin.com/eXYYUNn/")  #小顽童爷爷
    userInfo4 = getUserAll("https://v.douyin.com/eXFcnqd/")  #时尚奶奶团
    userInfo5 = getUserAll("https://v.douyin.com/eXYLVsR/")  #乐退族
    userInfo6 = getUserAll("https://v.douyin.com/eXFoHj2/")  #vk不省心大爷
    userInfo7 = getUserAll("https://v.douyin.com/eXFcTKQ/")  #孟小琦和爷爷
    userInfo8 = getUserAll("https://v.douyin.com/eXFK7wv/")  #姑妈有范儿
    userInfo9 = getUserAll("https://v.douyin.com/eXYFrmG/")  #雪瑞姑姑
    # userInfo10 = getUserAll("https://v.douyin.com/eXYVMd6/")  #爷爷等一下
    # print(userInfo10)
    userInfo11 = getUserAll("https://v.douyin.com/eXFcUWb/")  #康康和爷爷
    userInfo12 = getUserAll("https://v.douyin.com/eXFExyG/")  #淘气陈奶奶
    userInfo13 = getUserAll("https://v.douyin.com/eXFTxXr/")  #郎影和爷爷
    userInfo14 = getUserAll("https://v.douyin.com/eXFnKGj/")  #时尚奶奶
    userInfo15 = getUserAll("https://v.douyin.com/eXFnbKA/")  #国民姥姥
    userInfo16 = getUserAll("https://v.douyin.com/eXFtwFs/")  #最潮刘老头
    #
    #
    #
    videos_to_excel(userInfo1)
    videos_to_excel(userInfo2)
    videos_to_excel(userInfo3)
    videos_to_excel(userInfo4)
    videos_to_excel(userInfo5)
    videos_to_excel(userInfo6)
    videos_to_excel(userInfo7)
    videos_to_excel(userInfo8)
    videos_to_excel(userInfo9)
    # videos_to_excel(userInfo10)
    videos_to_excel(userInfo11)
    videos_to_excel(userInfo12)
    videos_to_excel(userInfo13)
    videos_to_excel(userInfo14)
    videos_to_excel(userInfo15)
    videos_to_excel(userInfo16)

    # videos_to_excel(userInfo)
    # videos_to_excel(userInfo1)
    # videos_to_excel(userInfo2)
    # videos_to_excel(userInfo3)
    # videos_to_excel(userInfo4)
    # videos_to_excel(userInfo5)
    # videos_to_excel(userInfo6)

    # print(json.dumps(userInfo))
